import os
import io
from fastapi import APIRouter, UploadFile, File, HTTPException
from typing import List
import PyPDF2
import docx
import openpyxl
from app.api.schemas import UploadResponse
from app.services.vector_store import vector_store

router = APIRouter()

UPLOAD_DIR = "./uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


def extract_text_from_pdf(file_content: bytes) -> str:
    text = ""
    try:
        pdf_file = io.BytesIO(file_content)
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error extracting PDF: {str(e)}")
    return text


def extract_text_from_docx(file_content: bytes) -> str:
    text = ""
    try:
        doc = docx.Document(io.BytesIO(file_content))
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error extracting DOCX: {str(e)}")
    return text


def extract_text_from_xlsx(file_content: bytes) -> str:
    text = ""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " | ".join([str(cell) if cell else "" for cell in row]) + "\n"
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error extracting XLSX: {str(e)}")
    return text


@router.post("/upload", response_model=UploadResponse)
async def upload_knowledge_base(files: List[UploadFile] = File(...)):
    if len(files) > 10:
        raise HTTPException(status_code=400, detail="Maximum 10 files allowed")

    total_added = 0

    for file in files:
        if file.size > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail=f"File {file.filename} too large (max 10MB)")

        content = await file.read()
        filename = file.filename.lower()

        if filename.endswith('.pdf'):
            text = extract_text_from_pdf(content)
        elif filename.endswith('.docx') or filename.endswith('.doc'):
            text = extract_text_from_docx(content)
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            text = extract_text_from_xlsx(content)
        elif filename.endswith('.txt'):
            text = content.decode('utf-8')
        else:
            continue

        if text.strip():
            vector_store.add_document(
                text=text,
                metadata={
                    "type": "uploaded_knowledge",
                    "source": file.filename,
                    "file_type": filename.split('.')[-1]
                }
            )
            total_added += 1

        save_path = os.path.join(UPLOAD_DIR, file.filename)
        with open(save_path, "wb") as f:
            f.write(content)

    return UploadResponse(
        filename=", ".join([f.filename for f in files]),
        status="success",
        documents_added=total_added,
        message=f"Successfully processed {total_added} documents"
    )


@router.get("/documents")
async def list_documents():
    docs = vector_store.get_all_documents()
    return {
        "total": len(docs),
        "documents": [
            {
                "id": doc["id"],
                "content_preview": doc["content"][:200] + "..." if len(doc["content"]) > 200 else doc["content"],
                "metadata": doc["metadata"]
            }
            for doc in docs
        ]
    }


@router.delete("/documents/{document_id}")
async def delete_document(document_id: str):
    try:
        vector_store.delete_document(document_id)
        return {"message": f"Document {document_id} deleted"}
    except Exception as e:
        raise HTTPException(status_code=404, detail="Document not found")
